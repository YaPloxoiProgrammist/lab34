import numpy as np
import cv2
from time import time
from openpyxl import load_workbook
import pandas

wb = load_workbook(filename='totalMarkClean.xlsx')
sheet = wb.worksheets[0]

# Считывание из excel столбца А
d = [row for row in sheet.iter_rows(min_row=1, max_col=1,
                                    max_row=sheet.max_row, values_only=True)]
Z = np.array(d, dtype=np.float32)
print(Z.shape, Z[0])

# define criteria, MAX_ITER, EPS, number of clusters(K) and apply kmeans()
criteria = (cv2.TERM_CRITERIA_EPS + cv2.TERM_CRITERIA_MAX_ITER, 10, 1)
K = 5
ret, label, center = cv2.kmeans(Z, K, None, criteria, 5, cv2.KMEANS_RANDOM_CENTERS)
print('label', label[:10])
print('label', label.shape)
print('ret', ret)
print('center', center.shape)
print('center', center)
print('===')
l = label.flatten()
print('l', l[:10])
print('l', l.shape)
for i in range(5):
    l1 = np.min(Z[l == i])
    l2 = np.max(Z[l == i])
    print('l', i, l1, l2)
for i in range(3408):
    # sheet.cell(row=i+1, column=2, value=l[i])
    sheet.cell(i + 1, 2, value=l[i])

wb.save('totMark.xls')